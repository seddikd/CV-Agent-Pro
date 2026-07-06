"""Génère un export Excel à partir de la table des candidats (SQLite)."""

from pathlib import Path
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

import web_db


# (en-tête, champ, largeur). champ « _date »/« _time » = dérivés de received_at.
COLUMNS = [
    ("ID", "id", 6),
    ("Date de réception", "_date", 14),
    ("Heure de réception", "_time", 10),
    ("Expéditeur", "expediteur", 30),
    ("Nom", "nom", 18),
    ("Prénom", "prenom", 18),
    ("Email", "email", 30),
    ("Téléphone", "telephone", 18),
    ("Poste recherché", "poste_recherche", 28),
    ("Spécialité", "specialite", 24),
    ("Wilaya / Ville", "wilaya", 16),
    ("Années d'expérience", "annees_experience", 12),
    ("Disponibilité", "disponibilite", 16),
    ("Salaire souhaité", "salaire_souhaite", 16),
    ("Niveau d'étude", "niveau_etude", 16),
    ("Diplôme le plus élevé", "diplome_plus_eleve", 28),
    ("Université", "universite", 26),
    ("Compétences techniques", "competences", 40),
    ("Logiciels", "logiciels", 26),
    ("Soft skills", "soft_skills", 26),
    ("Langues", "langues", 20),
    ("Certifications", "certifications", 30),
    ("Entreprises", "entreprises", 30),
    ("Permis", "permis", 12),
    ("Résumé du CV", "resume", 50),
    ("Nom du fichier PDF", "pdf_filename", 35),
    ("Chemin du PDF", "pdf_path", 50),
    ("Statut", "statut", 14),
    ("Commentaires RH", "commentaires", 30),
]


def _cell_value(c: dict, field: str):
    """Valeur d'une cellule pour un candidat, y compris les champs dérivés."""
    recv = c.get("received_at") or ""
    if field == "_date":
        return recv.partition("T")[0]
    if field == "_time":
        return recv.partition("T")[2][:8]
    v = c.get(field)
    return "" if v is None else v


def _build_workbook(rows: list[dict]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidats"
    ws.append([h for h, _, _ in COLUMNS])

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, (_, _, w) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    for c in rows:
        ws.append([_cell_value(c, field) for _, field, _ in COLUMNS])
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    return wb


def export_to_path(db_path: str, output_path: str) -> Path:
    rows = web_db.list_candidates(db_path, limit=100000)
    wb = _build_workbook(rows)
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def export_to_bytes(db_path: str) -> bytes:
    rows = web_db.list_candidates(db_path, limit=100000)
    wb = _build_workbook(rows)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
